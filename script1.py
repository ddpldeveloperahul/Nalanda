import os
import sys

# Add the Django project root directory ('ndis') to sys.path
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "ndis"))

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "ndis.settings")

import django
django.setup()

from openpyxl import load_workbook

from myapp.models import (
    State,
    District,
    SubDivision,
    Block,
)

# Excel file path
EXCEL_FILE = r"D:\Nalanda\Bihar_State_District_SubDivision_Block_Hierarchy.xlsx"

wb = load_workbook(EXCEL_FILE, data_only=True)
ws = wb.active

for row in ws.iter_rows(min_row=6, values_only=True):

    sr_no = row[0]
    state_name = row[1]
    division_name = row[2]    
    district_name = row[3]
    subdivision_name = row[4]
    block_name = row[5]

    if not state_name:
        continue

    # State
    state, _ = State.objects.get_or_create(
        name=state_name.strip()
    )

    # District
    district, _ = District.objects.get_or_create(
        state=state,
        name=district_name.strip()
    )

    # SubDivision
    subdivision, _ = SubDivision.objects.get_or_create(
        district=district,
        name=subdivision_name.strip()
    )

    # Block
    Block.objects.get_or_create(
        subdivision=subdivision,
        name=block_name.strip()
    )

print("✅ Bihar hierarchy imported successfully.")